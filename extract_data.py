import openpyxl
from pathlib import Path
# noinspection PyUnresolvedReferences
from xml.dom import minidom

# File where can find traceability links
trace_link_file = 'artifactWrappers.xmi'
trace_model_file = 'traceModel.xmi'
# Hazard file
hazard_file = 'mobstr-hazards.xlsx'
# Safety requirement file which was preprocessed to easier get content of each safety requirement
requirement_file = 'mobstr-requirements.xlsx'
# Original safety requirement file
req_file_original = 'mobstr-requirements_original.xlsx'


class Node:
    def __init__(self, id=None):
        self.id = id
        self.text = None
        self.target = []


class Failure(Node):
    pass


class Hazard(Node):
    pass


class SafetyGoal(Node):
    pass


class SafetyRequirement(Node):
    pass


def artifact_wrapper_handler():
    # Iterate the file where can find traceability links, to get the links of hazard and safety goal
    # This only apply to this specific dataset
    trace_link_doc = minidom.parse(trace_link_file)
    sub_graph_link = []
    for element in trace_link_doc.getElementsByTagName('artifacts'):
        if 'platform' in element.getAttribute('path'):
            # Split the path of hazard and safety goal in order to get the id of them
            # The third element of it is the id
            path_split = element.getAttribute('path').partition('row=')
            # Get the path of hazard and its related safety goal
            path = element.getAttribute('path').split('platform:/')[1].split('/mobstr-')[0]
            if 'hazard' in path:
                file = Path(hazard_file)
                file_object = openpyxl.load_workbook(file, data_only=True)
                sheet = file_object.active
                # Get the text of each hazard in the hazard .xlsx file
                for row in range(1, sheet.max_row + 1):
                    if sheet.cell(row=row, column=1).value == path_split[2]:
                        string = path_split[2] + ' ' + sheet.cell(row=row, column=2).value
                        # Save the text to a list
                        sub_graph_link.append(string)
            elif 'requirement' in path:
                file = Path(req_file_original)
                file_object = openpyxl.load_workbook(file, data_only=True)
                sheet = file_object.active
                # Get the text of each safety goal in the requirement .xlsx file
                for row in range(1, sheet.max_row + 1):
                    if sheet.cell(row=row, column=1).value == path_split[2]:
                        string = path_split[2] + ' ' + sheet.cell(row=row, column=3).value
                        # Save the text to the list
                        sub_graph_link.append(string)
        else:
            sub_graph_link.append(element.getAttribute('path'))

    return sub_graph_link


def add_failure():
    # Extract text of failure in traceModel.xmi and save it
    # Only apply to this specific dataset
    failure_list = []
    trace_model_doc = minidom.parse(trace_model_file)
    failure_id = 0
    for element in trace_model_doc.getElementsByTagName('traces'):
        failure_id += 1
        failure = Failure(failure_id)
        if 'Failure' in element.getAttribute('name'):
            failure.text = element.getAttribute('name').partition(': Failure ->')[0]
            failure_list.append(failure)
            add_hazard(element, failure)
    return failure_list


def add_hazard(element, failure):
    # Create a dictionary to save founded hazards
    dict_h = {}
    # Get the hazard and safety goal list
    sub_grpaph_link = artifact_wrapper_handler()
    # Iterate the traceModel.xmi to get the id of the hazards that each failure is related to
    for e1 in element.getElementsByTagName('target'):
        # Path of each related hazard
        path_related_hazard = e1.getAttribute('href')
        # Split the path in order to get the hazard id
        path_split = path_related_hazard.partition('@artifacts.')
        hazard_id = path_split[2]
        # Find the text of the hazard by the id in the list
        hazard_text = sub_grpaph_link[(int(hazard_id))]
        # If it's a new hazard, add it to the dictionary
        if hazard_id not in dict_h:
            hazard = Hazard(hazard_id)
            hazard.text = hazard_text
            current_hazard_id = hazard.text.split(' ')[0]
            dict_h[hazard_id] = hazard
        else:
            hazard = dict_h[hazard_id]
        failure.target.append(hazard)
        add_safety_goal(hazard, hazard_id, current_hazard_id)


def add_safety_goal(hazard, hazard_id, current_hazard_id):
    # Create a dictionary to save safety goal
    dict_s = {}
    safety_goal_id = hazard_id
    # If it's a new safety goal, add it to the dictionary
    if safety_goal_id not in dict_s:
        file = Path(req_file_original)
        obj = openpyxl.load_workbook(file, data_only=True)
        sheet = obj.active
        # Get text of related safety goal
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=2).value == current_hazard_id:
                safety_goal = SafetyGoal(safety_goal_id)
                safety_goal.text = sheet.cell(row=row, column=3).value
                dict_s[safety_goal_id] = safety_goal
                hazard.target.append(safety_goal)

    else:
        safety_goal = dict_s[safety_goal_id]
    add_safety_requirement(safety_goal, safety_goal_id, current_hazard_id)


def add_safety_requirement(safety_goal, safety_goal_id, current_hazard_id):
    dict_r = {}
    requirement_id = safety_goal_id
    if requirement_id not in dict_r:
        file = Path(requirement_file)
        obj = openpyxl.load_workbook(file, data_only=True)
        sheet = obj.active
        # Get text of all safety requirements
        for find_id in range(1, sheet.max_row):
            if sheet.cell(row=find_id, column=2).value is not None and current_hazard_id in sheet.cell(
                    row=find_id, column=2).value:
                id_row_number = find_id
                for row in range(id_row_number + 1, sheet.max_row):
                    if sheet.cell(row=row, column=1).value is not None and 'SR' in sheet.cell(row=row,
                                                                                              column=1).value:
                        requirement = SafetyRequirement(requirement_id)
                        requirement.text = sheet.cell(row=row, column=3).value.replace('The system shall ', '')
                        dict_r[requirement_id] = requirement
                        safety_goal.target.append(requirement)
                    if sheet.cell(row=row, column=1).value is not None and 'SG' in sheet.cell(row=row,
                                                                                              column=1).value:
                        break

    else:
        requirement = dict_r[requirement_id]


def print_everything(failure_list):
    # Print all traceability links to visualize
    for failure in failure_list:
        print()
        print("FAILURE")
        # print("Id")
        # print(failure.id)
        print("Failure Text: " + failure.text)
        print_hazards(failure.target)


def print_hazards(hazards):
    for hazard in hazards:
        print("HAZARD")
        # print("Id")
        # print(hazard.id)
        print("Hazard Text: " + hazard.text)
        print_safety_goals(hazard.target)


def print_safety_goals(safety_goals):
    for safety_goal in safety_goals:
        print("SAFETY GOAL")
        # print("id")
        # print(safety_goal.id)
        print("Safety goal Text: " + safety_goal.text)
        print_requirements(safety_goal.target)


def print_requirements(requirements):
    for requirement in requirements:
        print("REQUIREMENT")
        # print("id")
        # print(requirement.id)
        print("Safety requirement Text: ")
        print(requirement.text)

